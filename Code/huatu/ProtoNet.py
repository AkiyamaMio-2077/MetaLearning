from distutils.log import error
from tkinter import FALSE
from sklearn import datasets
import torch 
from torch.utils import data 
from train_utils import accuracy
from train_utils import seed_torch

from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import confusion_matrix
from sklearn.manifold import TSNE
import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns

import learn2learn as l2l
import time
import numpy as np
import pandas as pd 
from ProtoNet_Moudel import Net4CNN
import os
from init_utils import get_file , get_data_csv,normalization,sample_label_shuffle
from init_utils import weights_init2

from sklearn.metrics import recall_score,f1_score,accuracy_score,precision_score
from captum.attr import IntegratedGradients
import warnings

import sklearn.metrics as metrics
from sklearn.metrics import precision_recall_fscore_support
from sklearn.metrics import classification_report
from openpyxl import load_workbook

warnings.filterwarnings("ignore")
device = torch.device('cuda'if torch.cuda.is_available() else 'cpu')


class MetaDataset(torch.utils.data.Dataset):
    def __init__(self,file_path_train,file_path_Test,file_path_fine_tune,mode):
        if mode == 'train':
            self.data = pd.read_csv(file_path_train,index_col=False )
        if mode == 'valid':
            self.data = pd.read_csv(file_path_train,index_col=False)
        if mode == 'test':
            self.data = pd.read_csv(file_path_Test,index_col=False)
        if mode == 'fine_tune':
            self.data = pd.read_csv(file_path_fine_tune,index_col=False)

    def __len__(self):
        return len(self.data)

    def __getitem__(self,idx):

        label = self.data.iloc[idx, 0]  # 获取标签列
        name = self.data.iloc[idx, 1]  # 获取数据名称列
        sample = self.data.iloc[idx, 2:].values.astype(float)  # 获取数值列

        return sample,label,name


# csv_file_path_Train = 'F:\\data\\汇总\\FT_16\\drought\\0.csv'
# csv_file_path_Test = 'F:\\data\\汇总\\FT_16\\freeze\\0.csv'
csv_file_path_fine_tune = 'F:\\data\\汇总\\FT_16\\FT\\0.csv'
# meta_dataset = MetaDataset(csv_file_path_Train,csv_file_path_Test,mode = 'train')

class ProtoNet_learner(object):
    def __init__(self,ways):
        super().__init__()
        self.model = Net4CNN(hidden_size = 64,layers = 4,channels = 1).to(device)
        self.ways = ways

    def bulid_tasks(self,mode='train',ways = 10,shots = 5,num_tasks = 100,filter_labels = None ):
        meta_dataset = l2l.data.MetaDataset(MetaDataset(csv_file_path_Train,csv_file_path_Test,csv_file_path_fine_tune,mode = mode))
        new_ways = len(filter_labels) if filter_labels is not None else ways

        assert shots*new_ways <= meta_dataset.__len__()//ways*new_ways , "Reduce the number of shots!"

        tasks  = l2l.data.TaskDataset(meta_dataset,task_transforms = [
            l2l.data.transforms.FusedNWaysKShots(meta_dataset,new_ways,shots+1,filter_labels = filter_labels),
            l2l.data.transforms.LoadData(meta_dataset),
            # l2l.data.transforms.RemapLabels(meta_dataset,shuffle = True),
            l2l.data.transforms.ConsecutiveLabels(meta_dataset),
        ],num_tasks = num_tasks)

        return tasks
    def euclidean_metric(self,query_x, proto_x):
        """
        :param query_x: (n, d), N-example, P-dimension for each example; zq
        :param proto_x: (nc, d), M-Way, P-dimension for each example, but 1 example for each Way; z_proto
        :return: [n, nc]
        """

        query_x = query_x.unsqueeze(dim=1)  # [n, d]==>[n, 1, d]
        proto_x = proto_x.unsqueeze(dim=0)  # [nc, d]==>[1, nc, d]
        logits = -torch.pow(query_x - proto_x, 2).mean(dim=2)  # (n, nc)
        return logits


#这个fast_adapt 
    def fast_adapt(self,batch,learner,loss_fun,query_num,shots,ways):
        data ,labels,name = batch

        forclusterlist = labels        #重新映射标签
        le = LabelEncoder()
        le.fit(labels)
        labels = le.transform(labels)
        labels = torch.from_numpy(labels)

        data,labels = data.to(device).unsqueeze(1),labels.to(device)

         # Separate data into adaptation/evaluation sets
        sort = torch.sort(labels)
        data = data.squeeze(0)[sort.indices].squeeze(0)
        labels = labels.squeeze(0)[sort.indices].squeeze(0)

        support_indices = np.zeros(data.size(0),dtype = bool)
        selection = np.arange(ways)*(shots + query_num)   #0,shot+q ,2*(shot+q),3*(shot+q) ,.....   这部分可能后面要改，原本的代码时全部数据然后抽样的。
        
        for offset in range(shots):
            support_indices [selection + offset] = True    #0:shots ,(shot +q ):(shot+q+shots),....

        query_indices = torch.from_numpy(~support_indices) #shots:2*shots , (shot+q+shots):4*shots,....
        support_indices = torch.from_numpy(support_indices) #0:shots,(shot+q):(shot+q+shots),....
        
        embeddings = learner(data)
        
        support = embeddings[support_indices]
        support1 = support.reshape(ways,shots,-1)
        support = support1.mean(dim = 1) #(ways,dim)    
        query   = embeddings[query_indices]  #(n_query,dim)\

        labels  = labels[query_indices].long()

        logits = self.euclidean_metric(query,support)

        error = loss_fun(logits,labels)
        # acc = accuracy(logits,labels)


        #下面为做混淆矩阵做的改动，直接把accuracy 里面结构拿出来，得到预测值
        tragets = labels
        predictions = logits
        predictions = predictions.argmax(dim =-1).view(tragets.shape)
        acc = (predictions == tragets).sum().float()/ tragets.shape[0]
        predictions = predictions.cpu().detach().numpy()


        predictions = le.inverse_transform(predictions)
        labels = le.inverse_transform(labels.cpu().detach().numpy())

        tragets = labels
        recall = recall_score(tragets,predictions,average='weighted')
        F1 = f1_score(tragets,predictions,average='weighted')
        precision = precision_score(tragets,predictions,average='weighted')

        query = query.cpu().detach().numpy()

        return error,acc,recall,F1,precision,predictions,labels,query

    def train(self,save_path,shots):
        train_ways = valid_ways = self.ways
        query_num = 1
        print(f"{train_ways}--ways,{shots}--shots for training...")
        train_task = self.bulid_tasks('train',train_ways,shots,1000,None)
        valid_tasks = self.bulid_tasks('valid',valid_ways,shots,100,None)

        self.model.apply(weights_init2)
        optimizer = torch.optim.Adam(self.model.parameters(),lr=0.01)
        lr_scheduler = torch.optim.lr_scheduler.ExponentialLR(optimizer,gamma = 0.95)  #查一下
        loss_fun = torch.nn.CrossEntropyLoss()

        Epochs = 200
        Episodes = 100
        counter = 0
        
        for ep in range(Epochs):
            
            #training:
            t0 = time.time()
            self.model.train()
            meta_train_error = 0.0
            meta_train_accuracy = 0.0

            for epi in range(Episodes):
                batch = train_task.sample()
                loss,acc = self.fast_adapt(batch,self.model,loss_fun,query_num,shots,train_ways)

                meta_train_error += loss.item()
                meta_train_accuracy += acc.item()
                
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()

            lr_scheduler.step()
            t1 = time.time()
            print(f'*** Time /epoch {t1-t0:.4f} ***')
            print(f'epoch {ep+1}, train, loss: {meta_train_error/Episodes:.4f}, '
                  f'acc: {meta_train_accuracy/Episodes:.4f}')

            #validation
            self.model.eval()
            meta_valid_error = 0.0
            meta_valid_accuracy = 0.0
            for i,batch in enumerate(valid_tasks):
                with torch.no_grad():
                    loss,acc = self.fast_adapt(batch,self.model,loss_fun,query_num,shots,train_ways)
                meta_valid_error += loss.item()
                meta_valid_accuracy += acc.item()

            print(f'epoch {ep + 1}, validation, loss: {meta_valid_error / len(valid_tasks):.4f}, '
                    f'acc: {meta_valid_accuracy / len(valid_tasks):.4f}\n')

            counter +=1

            if (ep+1) >=200 and (ep+1)%2 == 0:
                if input('\n == Stop training? ==(y/n)\n').lower() == 'y':
                    new_save_path  = save_path +rf'_ep{ep + 1}'
                    self.model_save(new_save_path)
                    break
                elif input('\n == Save model? == (y/n)\n').lower() == 'y':
                    new_save_path = save_path + rf'_ep{ep + 1}'

    def test(self,load_path,shots):
        self.model.load_state_dict(torch.load(load_path))
        print(f'Load Model successfully from [{load_path}]...')

        test_ways = self.ways 
        query_num = 1
        print(f"{test_ways}--ways, {shots}--shots for testing ...") 
        test_tasks = self.bulid_tasks('test',test_ways,shots,50,None)
        loss_fun = torch.nn.CrossEntropyLoss()
        
        #validation
        self.model.eval()              
        meta_valid_error = 0.0
        meta_valid_accuracy = 0.0
        meta_F1 = 0.0
        meta_recall = 0.0
        meta_precision = 0.0

        t0 = time.time()

        y_pred =[]
        y_label = []
        y_embeddings = []

        for i,batch in enumerate(test_tasks):
            with torch.no_grad():
                loss,acc,recall,F1,precision,predications,Labels,embeddings  = self.fast_adapt(batch,self.model,loss_fun,query_num,shots,test_ways)

                y_pred1 = predications
                y_pred.extend(y_pred1)

                y_label1 = Labels
                y_label.extend(y_label1)


                y_embeddings1 = embeddings
                y_embeddings.extend(y_embeddings1)


            meta_valid_error += loss.item()
            meta_valid_accuracy += acc.item()
            meta_F1 += F1.item()
            meta_recall += recall.item()
            meta_precision += precision.item()

        t1 = time.time()
        print(f"*** Time for {len(test_tasks)} tasks {t1 - t0:.4f}(s)")
        print(f'Testing,loss:{meta_valid_error / len(test_tasks):.4f},\n'
              f'acc:{meta_valid_accuracy / len(test_tasks):.4f}\n'
              f'f1:{meta_F1 / len(test_tasks):.4f}\n'
              f'recall:{meta_recall / len(test_tasks):.4f}\n'
              f'precision:{meta_precision / len(test_tasks):.4f}')
        a = 1
    #此处是构建混淆矩阵
        # for i in range(len(y_label)):
        #     if y_label[i] != 0:
        #         y_label[i]= 1

        # for i in range(len(y_pred)):
        #     if y_pred[i] != 0:
        #         y_pred[i]= 1
        # total_samples = len(y_pred)
        # correct_predictions = sum(1 for y_pred, y_label in zip(y_pred,y_label ) if y_pred == y_label)
        # accuracy = correct_predictions / total_samples * 100
        # print(f"Accuracy: {accuracy}%")

        
        cm = confusion_matrix(y_label, y_pred)

        cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        # cm_normalized = cm.astype('float')*4
        cm_normalized = np.round(cm_normalized,4)
        cm_xtick = ['0','1','2','3','4','5']
        cm_ytick = ['0','1','2','3','4','5']




        test_label = y_label
        result = y_pred
        print("accuracy-score: {0:.4f}".format(accuracy_score(test_label,result)))
        print("F-score: {0:.4f}".format(f1_score(test_label,result,average='macro')))
        # precision_recall_fscore_support(test_label,result,average='macro')
        print(precision_recall_fscore_support(test_label,result,average='macro'))
        print(classification_report(test_label,result,digits = 4))

        diagonal_values = np.diag(cm_normalized)
        # 将对角线数据存储到pandas DataFrame
        df = pd.DataFrame({'Diagonal Data': diagonal_values})

        cm = confusion_matrix(y_label, y_pred)

        cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        # cm_normalized = cm.astype('float') 
        cm_normalized = np.round(cm_normalized,2)
        cm_xtick = ['0','1','2','3','4','5']
        cm_ytick = ['0','1','2','3','4','5']

        sns.heatmap(cm_normalized,fmt='g', cmap='Blues',annot=True,cbar=False,xticklabels=cm_xtick, yticklabels=cm_ytick)
        h=sns.heatmap(cm_normalized,fmt='g', cmap='Blues',annot=True,cbar=False) #画热力图，设置cbar=Falese
        cb=h.figure.colorbar(h.collections[0]) #显示colorbar
        plt.xlabel('True Label')
        plt.ylabel('Predict Label')
        # plt.show()

        # 指定要保存到的Excel文件
        # excel_file_path = r'C:\Users\soft\Desktop\123.xlsx'

        # sheet_name1='Way'+ str(way) +',shot'+str(shots) +''
        # sheet_name1 = 'Sheet1'

        # 将DataFrame保存到指定Excel文件
        # with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        #     writer.book = load_workbook(excel_file_path)


        #     # writer.sheets = {sheet.title: sheet for sheet in writer.book.worksheets}
        #     # fact_sheet=writer.sheets[sheet_name1] #type:openpyxl.worksheet.worksheet.Worksheet
            
        #     # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        #     df.to_excel(writer, sheet_name = sheet_name1, startrow=1, header=False, index=False)

        # # print(cm_normalized)
        # print(diagonal_values)

        # sns.heatmap(cm_normalized,fmt='g', cmap='Blues',annot=True,cbar=False,xticklabels=cm_xtick, yticklabels=cm_ytick)
        # h=sns.heatmap(cm_normalized,fmt='g', cmap='Blues',annot=True,cbar=False) #画热力图，设置cbar=Falese
        # cb=h.figure.colorbar(h.collections[0]) #显示colorbar
        # plt.xlabel('True Label')
        # plt.ylabel('Predict Label')
        # plt.show()


        # fig = plt.figure(figsize=(14,14))
        # fig.savefig(r'C:\Users\soft\Desktop\论文准备\输出图片.pdf',format='pdf',dpi=150)





        #此处是Tsne可视化
        tsne = TSNE(n_components=3,n_iter=2000)
        X_tsne_2d = tsne.fit_transform(y_embeddings)
        marker_list = ['.', ',', 'o', 'v', '^', '<', '>', '1', '2', '3', '4', '8', 's', 'p', 'P', '*', 'h', 'H', '+', 'x', 'X', 'D', 'd', '|', '_', 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        class_list = [0,1,2,3,4,5,6,7,8,9,10]
        # name = ['G10%','G20%','G30%','G40%','G50%','G60%','G70%','G80%','G90%','G100%','Ca100%']
        class_list = [0,1,2,3]
        name = ['0','1','2','3']
        n_class = len(class_list) # 测试集标签类别数
        palette = sns.hls_palette(n_class) # 配色方案
        sns.palplot(palette)
        
        # fig = plt.figure(figsize=(10, 8))
        # ax = fig.add_subplot(111, projection='3d')
        # y_label = np.array(y_label)
        # for idx,fruit in enumerate(class_list):
        #     # 获取颜色和点型
        #     color = palette[idx]
        #     marker = marker_list[idx%len(marker_list)]

        #     # 找到所有标注类别为当前类别的图像索引号
        #     indices = np.where(y_label == fruit)
        #     ax.scatter(X_tsne_2d[indices, 0], X_tsne_2d[indices, 1], X_tsne_2d[indices, 2],color=color, marker=marker, label=name[idx])
            
        # # ax.set_xlabel()
        # # ax.set_ylabel()
        # # ax.set_zlabel()
        # ax.legend(loc='best', title='Classes')
        # # ax.set_title('Sparse PCA of Data (3D)')

        fig = plt.figure(figsize=(14, 14))
        y_label = np.array(y_label)
        for idx,fruit in enumerate(class_list):
            # 获取颜色和点型
            color = palette[idx]
            marker = marker_list[idx%len(marker_list)]

            # 找到所有标注类别为当前类别的图像索引号
            indices = np.where(y_label == fruit)
            plt.scatter(X_tsne_2d[indices, 0], X_tsne_2d[indices, 1], color=color, marker=marker, label=fruit, s=150, clip_on=False)
            
        plt.legend(fontsize=16, markerscale=1, bbox_to_anchor=(1, 1))
        plt.xticks([])
        plt.yticks([])
        plt.show()





        plt.subplots_adjust(top=1, bottom=0, right=0.93, left=0, hspace=0, wspace=0)
        plt.show()
        # fig.savefig(r'C:\Users\soft\Desktop\论文准备\输出图片.svg',format='svg',dpi=150)

    def feature_importance2(self,load_path,shots):

        self.model.load_state_dict(torch.load(load_path))
        print(f'Load Model successfully from [{load_path}]...')
        test_ways = self.ways
        
        test_data = self.bulid_tasks('test',test_ways,shots,1000,None)
        loss_fun = torch.nn.CrossEntropyLoss()
        self.model.eval()

        X_train  = []
        X_label = []
        # test_data = self.build_data(mode='test', batch_size = 10)


        for batch in test_data:
            inputs, labels, _ = batch
            X_train.append(inputs)
            X_label.append(labels)

        # X_train = np.array(X_train)
        # X_train = np.vstack(X_train)

        # X_train = torch.from_numpy(X_train).type(torch.FloatTensor)

        ig = IntegratedGradients(self.model)
        inputs = inputs.to(device)
        inputs,labels = inputs.to(device).unsqueeze(1),labels.to(device)
        attr, delta = ig.attribute(inputs,target=1, return_convergence_delta=True)

        attr = attr.cpu().detach().numpy()

        feature_names = ["wave" + str(i) for i in range(21, 274)]
        importance= np.mean(attr,axis=0)
        print(importance)

        # 指定要保存到的Excel文件
        excel_file_path = r'C:\Users\soft\Desktop\123.xlsx'
        sheet_name1 = 'Sheet2'
        importance  = importance.flatten()
        importance = pd.DataFrame({'Diagonal Data': importance})


        # 将DataFrame保存到指定Excel文件
        # with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        #     writer.book = load_workbook(excel_file_path)
        #     importance.to_excel(writer, sheet_name = sheet_name1, startrow=1, header=False, index=False)



        # visualize_importances(feature_names, np.mean(attr, axis=0))




def visualize_importances(feature_names, importances, title="Average Feature Importances", plot=True, axis_title="Features"):
    print(title)
    for i in range(len(feature_names)):
        print(feature_names[i], ": ", '%.3f'%(importances[i]))
    x_pos = (np.arange(len(feature_names)))
    if plot:
        plt.figure(figsize=(12,6))
        plt.bar(x_pos, importances, align='center')
        plt.xticks(x_pos, feature_names, wrap=True)
        plt.xlabel(axis_title)
        plt.title(title)



if __name__ == "__main__":
    from train_utils import seed_torch
    seed_torch(2021)
    way  = 3 # 3 5 7 
    shot  = 3 #1 3 7 15 20 
    Net = ProtoNet_learner(ways = way)

    csv_file_path_Train = 'E:\\Milk\\averspectra\\1-D\\niu_tuo.csv'
    # csv_file_path_Train = 'F:\\data\\汇总\\FT_0\\freeze\\0.csv'
    csv_file_path_Test = 'E:\\Milk\\averspectra\\1-D\\test.csv'
    csv_file_path_fine_tune = 'E:\\Milk\\averspectra\\processing\\Train\\Fine_tuning\\FT_1.csv'

    path = r"E:\模型保存\milk\protonet\ProtoNet_milk_ways"+str(way)+"_shot "+ str(shot) +" "
    # Net.train(save_path = path ,shots = shot)
    # Net.feature_importance2(load_path=r"E:\模型保存\milk\protonet\ProtoNet_milk_ways"+str(way)+"_shot "+ str(shot) +" _ep200",shots=10 )

    Net.test(load_path=r"E:\模型保存\milk\protonet\ProtoNet_milk_ways"+str(way)+"_shot "+ str(shot) +" _ep200",shots=1)

    # for i in range(10):
    #     print(i,'\n')
    #     Net.test(load_path=r"E:\模型保存\milk\protonet\ProtoNet_milk_ways"+str(way)+"_shot "+ str(shot) +" _ep200",shots=i+1)
    #     Net.feature_importance2(load_path=r"E:\模型保存\milk\protonet\ProtoNet_milk_ways"+str(way)+"_shot "+ str(shot) +" _ep200",shots=i+1 )

    # # Net.fine_tune_test(load_path=r"E:\Paper\savepath\5shot\MAML\ProtoNet_FT0_shot2_ep200",shots=1)        

